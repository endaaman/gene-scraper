import openpyxl
import string
from Bio import Entrez
from Bio.Entrez import efetch, read

Entrez.email = "endaaman@eis.hokudai.ac.jp"
class database:
    def __init__(self):
        self.jcrissnlist = []
        self.jcrjtlist = []
        self.jcrisolist = []
        self.impactfactor = []

    def getexcellist(self):
        cell_word = []
        for cell in certainws['D']:
            self.jcrissnlist.append(cell.value)
        del self.jcrissnlist[0:2]
        print("*")

        for cell in certainws['B']:
            cell_word = cell.value
            if cell_word is None:
                continue
            else:
                cell_word = cell_word.translate(str.maketrans( '', '',string.punctuation))
                cell_word = cell_word.replace(" ", "")
                cell_word = cell_word.lower()
            self.jcrjtlist.append(cell_word)
        del self.jcrjtlist[0:1]
        print("*")

        for cell in certainws['C']:
            cell_word = cell.value
            if cell_word is None:
                continue
            else:
                cell_word = cell_word.translate(str.maketrans( '', '',string.punctuation))
                cell_word = cell_word.replace(" ", "")
                cell_word = cell_word.lower()
            self.jcrisolist.append(cell_word)
        del self.jcrisolist[0:1]
        print("*")

        for cell in certainws['F']:
            self.impactfactor.append(cell.value)
        del self.impactfactor[0:2]
        print("*")

        return (self.jcrissnlist, self.jcrjtlist, self.jcrisolist, self.impactfactor)

wb19 = openpyxl.load_workbook('2019.xlsx')
ws19 = wb19["2019"]
certainws = ws19
data19 = database()
data19.jcrissnlist = []
data19.jcrjtlist = []
data19.jcrisolist = []
data19.impactfactor = []
data19.getexcellist()

print("Loading data19 completed. First is: " + data19.jcrissnlist[0])
print("Loading data19 completed. First is: " + data19.jcrjtlist[0])
print("Loading data19 completed. First is: " + data19.jcrisolist[0])
print(data19.impactfactor[0])

wb18 = openpyxl.load_workbook('2018.xlsx')
ws18 = wb18["2018"]
certainws = ws18
data18 = database()
data18.jcrissnlist = []
data18.jcrjtlist = []
data18.jcrisolist = []
data18.impactfactor = []
data18.getexcellist()

print("Loading data18 completed. First is: " + data18.jcrissnlist[0])
print("Loading data18 completed. First is: " + data18.jcrjtlist[0])
print("Loading data18 completed. First is: " + data18.jcrisolist[0])
print(data18.impactfactor[0])

wb17 = openpyxl.load_workbook('2017.xlsx')
ws17 = wb17["2017"]
certainws = ws17
data17 = database()
data17.jcrissnlist = []
data17.jcrjtlist = []
data17.jcrisolist = []
data17.impactfactor = []
data17.getexcellist()

print("Loading data17 completed. First is: " + data17.jcrissnlist[0])
print("Loading data17 completed. First is: " + data17.jcrjtlist[0])
print("Loading data17 completed. First is: " + data17.jcrisolist[0])
print(data17.impactfactor[0])

wb16 = openpyxl.load_workbook('2016.xlsx')
ws16 = wb16["2016"]
certainws = ws16
data16 = database()
data16.jcrissnlist = []
data16.jcrjtlist = []
data16.jcrisolist = []
data16.impactfactor = []
data16.getexcellist()

print("Loading data16 completed. First is: " + data16.jcrissnlist[0])
print("Loading data16 completed. First is: " + data16.jcrjtlist[0])
print("Loading data16 completed. First is: " + data16.jcrisolist[0])
print(data16.impactfactor[0])

wb15 = openpyxl.load_workbook('2015.xlsx')
ws15 = wb15["2015"]
certainws = ws15
data15 = database()
data15.jcrissnlist = []
data15.jcrjtlist = []
data15.jcrisolist = []
data15.impactfactor = []
data15.getexcellist()

print("Loading data15 completed. First is: " + data15.jcrissnlist[0])
print("Loading data15 completed. First is: " + data15.jcrjtlist[0])
print("Loading data15 completed. First is: " + data15.jcrisolist[0])
print(data15.impactfactor[0])

wb14 = openpyxl.load_workbook('2014.xlsx')
ws14 = wb14["2014"]
certainws = ws14
data14 = database()
data14.jcrissnlist = []
data14.jcrjtlist = []
data14.jcrisolist = []
data14.impactfactor = []
data14.getexcellist()

print("Loading data14 completed. First is: " + data14.jcrissnlist[0])
print("Loading data14 completed. First is: " + data14.jcrjtlist[0])
print("Loading data14 completed. First is: " + data14.jcrisolist[0])
print(data14.impactfactor[0])

wb13 = openpyxl.load_workbook('2013.xlsx')
ws13 = wb13["2013"]
certainws = ws13
data13 = database()
data13.jcrissnlist = []
data13.jcrjtlist = []
data13.jcrisolist = []
data13.impactfactor = []
data13.getexcellist()

print("Loading data13 completed. First is: " + data13.jcrissnlist[0])
print("Loading data13 completed. First is: " + data13.jcrjtlist[0])
print("Loading data13 completed. First is: " + data13.jcrisolist[0])
print(data13.impactfactor[0])

wb12 = openpyxl.load_workbook('2012.xlsx')
ws12 = wb12["2012"]
certainws = ws12
data12 = database()
data12.jcrissnlist = []
data12.jcrjtlist = []
data12.jcrisolist = []
data12.impactfactor = []
data12.getexcellist()

print("Loading data12 completed. First is: " + data12.jcrissnlist[0])
print("Loading data12 completed. First is: " + data12.jcrjtlist[0])
print("Loading data12 completed. First is: " + data12.jcrisolist[0])
print(data12.impactfactor[0])

wb11 = openpyxl.load_workbook('2011.xlsx')
ws11 = wb11["2011"]
certainws = ws11
data11 = database()
data11.jcrissnlist = []
data11.jcrjtlist = []
data11.jcrisolist = []
data11.impactfactor = []
data11.getexcellist()

print("Loading data11 completed. First is: " + data11.jcrissnlist[0])
print("Loading data11 completed. First is: " + data11.jcrjtlist[0])
print("Loading data11 completed. First is: " + data11.jcrisolist[0])
print(data11.impactfactor[0])

wb10 = openpyxl.load_workbook('2010.xlsx')
ws10 = wb10["2010"]
certainws = ws10
data10 = database()
data10.jcrissnlist = []
data10.jcrjtlist = []
data10.jcrisolist = []
data10.impactfactor = []
data10.getexcellist()

print("Loading data10 completed. First is: " + data10.jcrissnlist[0])
print("Loading data10 completed. First is: " + data10.jcrjtlist[0])
print("Loading data10 completed. First is: " + data10.jcrisolist[0])
print(data10.impactfactor[0])

wb09 = openpyxl.load_workbook('2009.xlsx')
ws09 = wb09["2009"]
certainws = ws09
data09 = database()
data09.jcrissnlist = []
data09.jcrjtlist = []
data09.jcrisolist = []
data09.impactfactor = []
data09.getexcellist()

print("Loading data09 completed. First is: " + data09.jcrissnlist[0])
print("Loading data09 completed. First is: " + data09.jcrjtlist[0])
print("Loading data09 completed. First is: " + data09.jcrisolist[0])
print(data09.impactfactor[0])

wb08 = openpyxl.load_workbook('2008.xlsx')
ws08 = wb08["2008"]
certainws = ws08
data08 = database()
data08.jcrissnlist = []
data08.jcrjtlist = []
data08.jcrisolist = []
data08.impactfactor = []
data08.getexcellist()

print("Loading data08 completed. First is: " + data08.jcrissnlist[0])
print("Loading data08 completed. First is: " + data08.jcrjtlist[0])
print("Loading data08 completed. First is: " + data08.jcrisolist[0])
print(data08.impactfactor[0])

wb07 = openpyxl.load_workbook('2007.xlsx')
ws07 = wb07["2007"]
certainws = ws07
data07 = database()
data07.jcrissnlist = []
data07.jcrjtlist = []
data07.jcrisolist = []
data07.impactfactor = []
data07.getexcellist()

print("Loading data07 completed. First is: " + data07.jcrissnlist[0])
print("Loading data07 completed. First is: " + data07.jcrjtlist[0])
print("Loading data07 completed. First is: " + data07.jcrisolist[0])
print(data07.impactfactor[0])

wb06 = openpyxl.load_workbook('2006.xlsx')
ws06 = wb06["2006"]
certainws = ws06
data06 = database()
data06.jcrissnlist = []
data06.jcrjtlist = []
data06.jcrisolist = []
data06.impactfactor = []
data06.getexcellist()

print("Loading data06 completed. First is: " + data06.jcrissnlist[0])
print("Loading data06 completed. First is: " + data06.jcrjtlist[0])
print("Loading data06 completed. First is: " + data06.jcrisolist[0])
print(data06.impactfactor[0])

wb05 = openpyxl.load_workbook('2005.xlsx')
ws05 = wb05["2005"]
certainws = ws05
data05 = database()
data05.jcrissnlist = []
data05.jcrjtlist = []
data05.jcrisolist = []
data05.impactfactor = []
data05.getexcellist()

print("Loading data05 completed. First is: " + data05.jcrissnlist[0])
print("Loading data05 completed. First is: " + data05.jcrjtlist[0])
print("Loading data05 completed. First is: " + data05.jcrisolist[0])
print(data05.impactfactor[0])

wb04 = openpyxl.load_workbook('2004.xlsx')
ws04 = wb04["2004"]
certainws = ws04
data04 = database()
data04.jcrissnlist = []
data04.jcrjtlist = []
data04.jcrisolist = []
data04.impactfactor = []
data04.getexcellist()

print("Loading data04 completed. First is: " + data04.jcrissnlist[0])
print("Loading data04 completed. First is: " + data04.jcrjtlist[0])
print("Loading data04 completed. First is: " + data04.jcrisolist[0])
print(data04.impactfactor[0])

wb03 = openpyxl.load_workbook('2003.xlsx')
ws03 = wb03["2003"]
certainws = ws03
data03 = database()
data03.jcrissnlist = []
data03.jcrjtlist = []
data03.jcrisolist = []
data03.impactfactor = []
data03.getexcellist()

print("Loading data03 completed. First is: " + data03.jcrissnlist[0])
print("Loading data03 completed. First is: " + data03.jcrjtlist[0])
print("Loading data03 completed. First is: " + data03.jcrisolist[0])
print(data03.impactfactor[0])

wb02 = openpyxl.load_workbook('2002.xlsx')
ws02 = wb02["2002"]
certainws = ws02
data02 = database()
data02.jcrissnlist = []
data02.jcrjtlist = []
data02.jcrisolist = []
data02.impactfactor = []
data02.getexcellist()

print("Loading data02 completed. First is: " + data02.jcrissnlist[0])
print("Loading data02 completed. First is: " + data02.jcrjtlist[0])
print("Loading data02 completed. First is: " + data02.jcrisolist[0])
print(data02.impactfactor[0])

wb01 = openpyxl.load_workbook('2001.xlsx')
ws01 = wb01["2001"]
certainws = ws01
data01 = database()
data01.jcrissnlist = []
data01.jcrjtlist = []
data01.jcrisolist = []
data01.impactfactor = []
data01.getexcellist()

print("Loading data01 completed. First is: " + data01.jcrissnlist[0])
print("Loading data01 completed. First is: " + data01.jcrjtlist[0])
print("Loading data01 completed. First is: " + data01.jcrisolist[0])
print(data01.impactfactor[0])

print("Loading database completed.")



wbimport = openpyxl.load_workbook('oda2up.xlsx')
wsimport = wbimport["up2,w.genesymbol"]
pregenelist = []
genelist = []

for cell in wsimport['C']:
    pregenelist.append(cell.value)
del pregenelist[0]

pregenelist = [e for e in pregenelist if e is not None]

genelist = sorted(set(pregenelist), key=pregenelist.index)

print("Loading genelist completed.")
print(genelist)
print(len(genelist))


print("word?")
searchword = input()
print("YYYY/MM/DD")
searchmindate = input()
searchmaxdate = input()



for certaingene in genelist:
    genename = certaingene + " AND " + searchword
    print(genename)
    handle = Entrez.esearch(db="pubmed", term=genename, mindate=searchmindate, maxdate=searchmaxdate, retmax=100000)
    record = Entrez.read(handle)
    countID = record["Count"]
    print("Hit article count: " + countID)

print("Remove word?")
removeword = input()
genelist.remove(removeword)
print(genelist)

print(len(genelist))

allgenename = []
allsumimpactfactor = []
allarticlecount = []
allIdList = []
alltitlelist = []
allyearlist = []
alljtlist = []
allissnlist = []
allisolist = []
allimpactfactor = []
allifsituationlist = []
allifyearlist = []
allIdLink = []

for certaingene in genelist:
    if certaingene == "Gene_Name":
        continue
    genename = certaingene + " AND " + searchword
    allgenename.append(genename)
    print(genename)
    handle = Entrez.esearch(db="pubmed", term=genename, mindate=searchmindate, maxdate=searchmaxdate, retmax=100000)
    record = Entrez.read(handle)
    countID = record["Count"]
    IdList = record["IdList"]
    allIdList.extend(IdList)
    print("Hit article count: " + countID)

    if countID == "0":
        allarticlecount.append("0")
        alltitlelist.append("")
        allyearlist.append("")
        allissnlist.append("")
        alljtlist.append("")
        allisolist.append("")
        allimpactfactor.append("")
        allifsituationlist.append("")
        allifyearlist.append("")
        allsumimpactfactor.append("")
        allIdList.append("")
        allIdLink.append("")
        continue
    titlelist = []
    yearlist = []
    jtlist = []
    issnlist = []
    isolist = []
    getimpactfactor = []
    articlecount = 0
    ifsituationlist = []
    ifyearlist = []
    sumimpactfactor = 0
    repeatcount = 0

    for ID in IdList:
        allIdLink.append("https://pubmed.ncbi.nlm.nih.gov/" + ID + "/")
        repeatcount += 1

        if repeatcount == 1:
            pass
        else:
            allgenename.append("")
        issn = []
        handle = efetch(db="pubmed", id=ID, retmode="xml")
        xml_data = read(handle)
        articledata = xml_data["PubmedArticle"][0]["MedlineCitation"]["Article"]

        try:
            titledata = articledata["ArticleTitle"]
        except:
            titledata = []
            print("False:titledata")
        strtitledata = str(titledata)
        titlelist.append(strtitledata)

        try:
            yeardata = articledata["Journal"]["JournalIssue"]["PubDate"]["MedlineDate"]
        except:
            try:
                yeardata = articledata["Journal"]["JournalIssue"]["PubDate"]["Year"]
            except:
                yeardata = []
                print("False:yeardata")
        stryeardata = str(yeardata)
        stryeardata = stryeardata[:4]
        yearlist.append(stryeardata)

        try:
            issndata = articledata["Journal"]["ISSN"]
        except:
            issndata = []
            print("False:issndata")
        strissndata = str(issndata)
        issnlist.append(strissndata)

        try:
            jtdata = articledata["Journal"]["Title"]
        except:
            jtdata = []
            print("False:jtdata")
        strjtdata = str(jtdata)
        jtlist.append(strjtdata)

        strjtdata = strjtdata.translate(str.maketrans( '', '',string.punctuation))
        strjtdata = strjtdata.replace(" ", "")
        strjtdata = strjtdata.lower()

        try:
            isodata = articledata["Journal"]["ISOAbbreviation"]
        except:
            isodata = []
            print("False:isodata")
        strisodata = str(isodata)

        isolist.append(strisodata)

        strisodata = strisodata.translate(str.maketrans( '', '',string.punctuation))
        strisodata = strisodata.replace(" ", "")
        strisodata = strisodata.lower()

        jcrissnlist = []
        jcrjtlist = []
        jcrisolist = []
        impactfactor = []
        imp = [0]
        getsituation = ()
        getyear = ()
        for year in ["19", "18", "17", "16", "15", "14", "13", "12", "11", "10", "09", "08", "07", "06", "05", "04", "03", "02", "01"]:
            if year == "19":
                jcrissnlist = data19.jcrissnlist
                jcrjtlist = data19.jcrjtlist
                jcrisolist = data19.jcrisolist
                impactfactor = data19.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "18" and imp == 0:
                jcrissnlist = data18.jcrissnlist
                jcrjtlist = data18.jcrjtlist
                jcrisolist = data18.jcrisolist
                impactfactor = data18.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "17" and imp == 0:
                jcrissnlist = data17.jcrissnlist
                jcrjtlist = data17.jcrjtlist
                jcrisolist = data17.jcrisolist
                impactfactor = data17.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "16" and imp == 0:
                jcrissnlist = data16.jcrissnlist
                jcrjtlist = data16.jcrjtlist
                jcrisolist = data16.jcrisolist
                impactfactor = data16.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "15" and imp == 0:
                jcrissnlist = data15.jcrissnlist
                jcrjtlist = data15.jcrjtlist
                jcrisolist = data15.jcrisolist
                impactfactor = data15.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "14" and imp == 0:
                jcrissnlist = data14.jcrissnlist
                jcrjtlist = data14.jcrjtlist
                jcrisolist = data14.jcrisolist
                impactfactor = data14.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "13" and imp == 0:
                jcrissnlist = data13.jcrissnlist
                jcrjtlist = data13.jcrjtlist
                jcrisolist = data13.jcrisolist
                impactfactor = data13.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "12" and imp == 0:
                jcrissnlist = data12.jcrissnlist
                jcrjtlist = data12.jcrjtlist
                jcrisolist = data12.jcrisolist
                impactfactor = data12.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "11" and imp == 0:
                jcrissnlist = data11.jcrissnlist
                jcrjtlist = data11.jcrjtlist
                jcrisolist = data11.jcrisolist
                impactfactor = data11.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "10" and imp == 0:
                jcrissnlist = data10.jcrissnlist
                jcrjtlist = data10.jcrjtlist
                jcrisolist = data10.jcrisolist
                impactfactor = data10.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "09" and imp == 0:
                jcrissnlist = data09.jcrissnlist
                jcrjtlist = data09.jcrjtlist
                jcrisolist = data09.jcrisolist
                impactfactor = data09.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "08" and imp == 0:
                jcrissnlist = data08.jcrissnlist
                jcrjtlist = data08.jcrjtlist
                jcrisolist = data08.jcrisolist
                impactfactor = data08.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "07" and imp == 0:
                jcrissnlist = data07.jcrissnlist
                jcrjtlist = data07.jcrjtlist
                jcrisolist = data07.jcrisolist
                impactfactor = data07.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "06" and imp == 0:
                jcrissnlist = data06.jcrissnlist
                jcrjtlist = data06.jcrjtlist
                jcrisolist = data06.jcrisolist
                impactfactor = data06.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "05" and imp == 0:
                jcrissnlist = data05.jcrissnlist
                jcrjtlist = data05.jcrjtlist
                jcrisolist = data05.jcrisolist
                impactfactor = data05.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "04" and imp == 0:
                jcrissnlist = data04.jcrissnlist
                jcrjtlist = data04.jcrjtlist
                jcrisolist = data04.jcrisolist
                impactfactor = data04.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "03" and imp == 0:
                jcrissnlist = data03.jcrissnlist
                jcrjtlist = data03.jcrjtlist
                jcrisolist = data03.jcrisolist
                impactfactor = data03.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "02" and imp == 0:
                jcrissnlist = data02.jcrissnlist
                jcrjtlist = data02.jcrjtlist
                jcrisolist = data02.jcrisolist
                impactfactor = data02.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

            elif year == "01" and imp == 0:
                jcrissnlist = data01.jcrissnlist
                jcrjtlist = data01.jcrjtlist
                jcrisolist = data01.jcrisolist
                impactfactor = data01.impactfactor
                try:
                    imp = jcrissnlist.index(strissndata)
                    getsituation = "issnhit"
                    getyear = year
                except:
                    try:
                        imp = jcrjtlist.index(strjtdata)
                        getsituation = "jthit"
                        getyear = year
                    except:
                        try:
                            imp = jcrisolist.index(strisodata)
                            getsituation = "isohit"
                            getyear = year
                        except:
                            imp = 0

        if imp == 0:
            getsituation = "N/A"
            getyear = "N/A"
            imp = "N/A"
        else:
            imp = impactfactor[imp]
            if imp == "Not Available":
                getsituation = "N/A"
                getyear = "N/A"
                imp = "N/A"
            else:
                sumimpactfactor = float(sumimpactfactor) + float(imp)
                sumimpactfactor = format(sumimpactfactor, '.3f')

        getimpactfactor.append(imp)
        ifsituationlist.append(getsituation)
        ifyearlist.append(getyear)
        print(getsituation)
        print(getyear)
        print(imp)

    alltitlelist.extend(titlelist)
    allyearlist.extend(yearlist)
    allissnlist.extend(issnlist)
    alljtlist.extend(jtlist)
    allisolist.extend(isolist)
    allimpactfactor.extend(getimpactfactor)
    allifsituationlist.extend(ifsituationlist)
    allifyearlist.extend(ifyearlist)
    allsumimpactfactor.append(sumimpactfactor)
    allarticlecount.append(countID)
    for x in range(repeatcount - 1):
        allsumimpactfactor.append("")
        allarticlecount.append("")
    print(allgenename)
    print(allarticlecount)
    print(alltitlelist)
    print(allyearlist)
    print(allissnlist)
    print(alljtlist)
    print(allisolist)
    print(allimpactfactor)
    print(allifsituationlist)
    print(allifyearlist)
    print(allsumimpactfactor)
    print(allIdList)
    print(allIdLink)

import winsound
duration = 1000
freq = 440
winsound.Beep(freq, duration)

print("Searching completed.")

allgenename.insert(0, "Search word")
allarticlecount.insert(0, "Hit No.")
alltitlelist.insert(0, "Article title")
allyearlist.insert(0, "Year")
alljtlist.insert(0, "Full journal title")
allisolist.insert(0, "J. abbrev.")
allimpactfactor.insert(0, "Impact factor")
allsumimpactfactor.insert(0, "Total IF")
allIdList.insert(0, "PubMed ID")
allIdLink.insert(0, "URL")


import pandas as pd
import numpy as np
matrix = [allgenename, allarticlecount, alltitlelist, allyearlist, alljtlist, allisolist, allimpactfactor, allsumimpactfactor, allIdList, allIdLink]

tr = []
for i in range(len(alltitlelist)):
    tr_row = []
    for vector in matrix:
        tr_row.append(vector[i])
    tr.append(tr_row)

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb=Workbook()
ws=wb.active

for row in tr:
    ws.append(row)
wb.save(filename="genedata.xlsx")
